"""보고서 JSON → Excel(.xlsx) 변환."""
import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GREEN = "2E7D5B"
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="맑은 고딕", size=10)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14, color=NAVY)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)


def _safe_sheet_name(name: str, used: set) -> str:
    name = re.sub(r"[\\/*?:\[\]]", " ", name).strip()[:28] or "Sheet"
    base, i = name, 2
    while name in used:
        name = f"{base[:25]}_{i}"
        i += 1
    used.add(name)
    return name


def _write_header_row(ws, row, values, fill_color):
    fill = PatternFill("solid", fgColor=fill_color)
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=str(v))
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.border = THIN
        cell.alignment = WRAP


def _autofit(ws, max_width=60):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def build_xlsx(report: dict) -> bytes:
    wb = Workbook()
    used_names = set()

    # 1) 요약 시트
    ws = wb.active
    ws.title = _safe_sheet_name("요약", used_names)
    ws.cell(row=1, column=1, value=report.get("title", "조사 보고서")).font = TITLE_FONT
    ws.cell(row=2, column=1, value=report.get("subtitle", "")).font = BODY_FONT

    row = 4
    _write_header_row(ws, row, ["Executive Summary"], NAVY)
    row += 1
    cell = ws.cell(row=row, column=1, value=report.get("executive_summary", ""))
    cell.font = BODY_FONT
    cell.alignment = WRAP
    ws.row_dimensions[row].height = 120
    ws.column_dimensions["A"].width = 100

    row += 2
    _write_header_row(ws, row, ["핵심 발견사항"], NAVY)
    for item in report.get("key_findings", []):
        row += 1
        cell = ws.cell(row=row, column=1, value="• " + str(item))
        cell.font = BODY_FONT
        cell.alignment = WRAP

    row += 2
    _write_header_row(ws, row, ["제언"], NAVY)
    for item in report.get("recommendations", []):
        row += 1
        cell = ws.cell(row=row, column=1, value="• " + str(item))
        cell.font = BODY_FONT
        cell.alignment = WRAP

    # 2) 본문 시트
    ws = wb.create_sheet(_safe_sheet_name("본문", used_names))
    _write_header_row(ws, 1, ["섹션", "내용", "요점"], NAVY)
    for r, sec in enumerate(report.get("sections", []), start=2):
        ws.cell(row=r, column=1, value=sec.get("heading", "")).font = Font(
            name="맑은 고딕", bold=True, size=10
        )
        c2 = ws.cell(row=r, column=2, value=sec.get("content", ""))
        c3 = ws.cell(row=r, column=3, value="\n".join(sec.get("bullets") or []))
        for cell in (ws.cell(row=r, column=1), c2, c3):
            cell.alignment = WRAP
            cell.border = THIN
        c2.font = BODY_FONT
        c3.font = BODY_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 60

    # 3) 데이터 표 시트 (표마다 하나)
    for table_data in report.get("data_tables", []):
        ws = wb.create_sheet(
            _safe_sheet_name(table_data.get("title", "데이터"), used_names)
        )
        ws.cell(row=1, column=1, value=table_data.get("title", "")).font = TITLE_FONT
        headers = table_data.get("headers", [])
        if headers:
            _write_header_row(ws, 3, headers, GREEN)
            for r, row_vals in enumerate(table_data.get("rows", []), start=4):
                for c in range(len(headers)):
                    val = str(row_vals[c]) if c < len(row_vals) else ""
                    cell = ws.cell(row=r, column=c + 1, value=val)
                    cell.font = BODY_FONT
                    cell.border = THIN
                    cell.alignment = WRAP
        _autofit(ws)

    # 4) 출처 시트
    if report.get("sources"):
        ws = wb.create_sheet(_safe_sheet_name("출처", used_names))
        _write_header_row(ws, 1, ["제목", "URL"], NAVY)
        for r, s in enumerate(report["sources"], start=2):
            ws.cell(row=r, column=1, value=s.get("title", "")).font = BODY_FONT
            cell = ws.cell(row=r, column=2, value=s.get("url", ""))
            cell.font = Font(name="맑은 고딕", size=10, color="0563C1", underline="single")
            if s.get("url", "").startswith("http"):
                cell.hyperlink = s["url"]
        _autofit(ws, max_width=80)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
